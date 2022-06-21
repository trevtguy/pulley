from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import os
import glob
import shutil

def findCell(file, datalist, pdfname, dst_path):
  
    listCounter = 0
    #opens both the tabulated data work book and the current dataset file
    toWorkbook = load_workbook(filename="tabulated_data.xlsx")
    workbook = load_workbook(filename=file,data_only=True)
    #sets the workbook sheets to active
    sheet = workbook.active
    tosheet = toWorkbook.active
    #finds the last row in the tabulated data workbook
    tolast_empty_row = tosheet.max_row + 1
    
    for x in datalist:
        #searches through the dataset file to find the appropriate value and add to the
        #tabulated data workbook
        cellCounter = 1
        for col in sheet.iter_cols(min_row=1, max_col=1, max_row=400, values_only=True):
            for idcell in col:
                if idcell == datalist[listCounter]:
                    counterstr = str(cellCounter)
                    val = sheet["B" + counterstr].value
                    if x == "Job Number":
                        # print(val[0])
                        if val[0] != "G":
                            valint = int(val)
                            if valint < 7000:
                                val = "G" + val
                                print(val)
                        # val = str(val)
                        # val = "12345" + val
                    c = tosheet.cell(row=tolast_empty_row, column=listCounter+1, value=val )
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    #if hyperlink is available it is added to the tabulated data file
                    if pdfname != None:
                        pdfpath = '=HYPERLINK("' + dst_path + '\\' + pdfname +'","GA DWG")'
                        hlink = tosheet.cell(row=tolast_empty_row, column=2, value=pdfpath )
                        hlink.font = Font(color = '0000FF', underline= 'single')
                    toWorkbook.save(filename="tabulated_data.xlsx")
                cellCounter += 1
            listCounter += 1
    toWorkbook.close() 

def copyDrawing(file, dst_path):
    rev = 8
    while rev >= 0:
        #sets up the standard file path for pdf files in each pulley job folder
        pdfpath = file.rsplit('\\', 2)
        pdfpathfull = pdfpath[0] + '\pdf files\\' + pdfpath[1] + '_' + str(rev) +'.pdf'
        pdfname = pdfpath[1] + '_' + str(rev) +'.pdf'
        rev -= 1
        try:
            #if the file exists the file name is returned or else it is 'skipped'
            file_exists = os.path.exists(pdfpathfull)
            if file_exists:
                shutil.copy(pdfpathfull, dst_path)
                return(pdfname)
        except:
            pass
        # DIFFERENCE



#file path to where the drawings will be saved
dst_path = r"C:\Users\Trevor\hello\drawings"
#List of data to be extracted from each dataset
#link leaves a spare column for the hyperlink to the pdf drawing
datalist = [
    "Part Number","link", "Job Number", "SHELL_LENGTH_FACE", "SHELL_DIAMETER_AFTER_MACHINING", "Company",
    "BEARINGS"
    ]
#creates a list of the files in the current working directory
cwd = os.getcwd()
#searches through the cwd list to find excel files that start with 750
for file in glob.glob(cwd+ "/**/750-*xlsx", recursive=True):
    # print(file)
    #if a pdf file is available copies it to the drawings folder
    pdfname = copyDrawing(file, dst_path)
    #transfers data from each pulley dataset and puts it in the tabulated data file
    findCell(file, datalist, pdfname, dst_path)
